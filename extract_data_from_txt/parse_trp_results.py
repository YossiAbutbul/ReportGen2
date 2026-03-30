from pathlib import Path
import re
import sys
import os
import shutil
import subprocess
from copy import copy

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parent

DEFAULT_ROOT_DIR = r"C:\Users\yossefa\OneDrive - Arad Technologies Ltd\Desktop\Work\2026\LoRa Columbia\LoRa Columbia"
DEFAULT_TEMPLATE_PATH = REPO_ROOT / "public" / "report-generator-template.xlsx"
DEFAULT_EXCEL_OUTPUT = SCRIPT_DIR / "trp_summary.xlsx"

FREQ_RE = re.compile(r"Test Frequency:\s*([+-]?\d+(?:\.\d+)?)\s*MHz", re.IGNORECASE)
TRP_RE = re.compile(r"Calculated TRP\s*=\s*([+-]?\d+(?:\.\d+)?)\s*dBm", re.IGNORECASE)

TABLE_ROW_RE = re.compile(
    r"^\s*"
    r"([+-]?\d+(?:\.\d+)?)\s+"
    r"([+-]?\d+(?:\.\d+)?)\s+"
    r"([+-]?\d+(?:\.\d+)?)\s+"
    r"([+-]?\d+(?:\.\d+)?)\s*$"
)


def extract_frequency(text: str, file_name: str) -> float | None:
    match = FREQ_RE.search(text)
    if match:
        return float(match.group(1))

    fallback = re.search(r"([+-]?\d+(?:\.\d+)?)\s*TRP", file_name, re.IGNORECASE)
    if fallback:
        return float(fallback.group(1))

    return None


def extract_trp(text: str) -> float | None:
    match = TRP_RE.search(text)
    if match:
        return float(match.group(1))
    return None


def extract_max_vpol(text: str) -> float | None:
    lines = text.splitlines()
    in_results_table = False
    vpol_values = []

    for line in lines:
        if "*******  Test Data Results  ********" in line:
            in_results_table = True
            continue

        if not in_results_table:
            continue

        row_match = TABLE_ROW_RE.match(line)
        if row_match:
            vpol_values.append(float(row_match.group(4)))

    if not vpol_values:
        return None

    return max(vpol_values)


def parse_file(txt_path: Path, unit_id: str) -> dict:
    text = txt_path.read_text(encoding="utf-8", errors="ignore")

    return {
        "unit_id": unit_id,
        "file_name": txt_path.name,
        "frequency_mhz": extract_frequency(text, txt_path.name),
        "trp_dbm": extract_trp(text),
        "max_peak_vpol_dbm": extract_max_vpol(text),
    }


def scan_root(root_dir: Path) -> list[dict]:
    results = []

    for unit_dir in sorted(root_dir.iterdir()):
        if not unit_dir.is_dir():
            continue

        unit_id = unit_dir.name

        for txt_file in sorted(unit_dir.glob("*.txt")):
            try:
                results.append(parse_file(txt_file, unit_id))
            except Exception as exc:
                results.append({
                    "unit_id": unit_id,
                    "file_name": txt_file.name,
                    "frequency_mhz": None,
                    "trp_dbm": None,
                    "max_peak_vpol_dbm": None,
                    "error": str(exc),
                })

    return results


def sort_rows_for_excel(rows: list[dict]) -> list[dict]:
    return sorted(
        rows,
        key=lambda row: (
            str(row.get("unit_id") or ""),
            float(row.get("frequency_mhz") or float("inf")),
            str(row.get("file_name") or ""),
        ),
    )


def copy_template_file(template_path: Path, output_file: Path) -> None:
    output_file.parent.mkdir(parents=True, exist_ok=True)

    try:
        shutil.copy2(template_path, output_file)
        return
    except PermissionError:
        pass

    command = [
        "powershell",
        "-NoProfile",
        "-Command",
        f"Copy-Item -LiteralPath '{template_path}' -Destination '{output_file}' -Force",
    ]
    subprocess.run(command, check=True)


def copy_row_format(worksheet, source_row: int, target_row: int, max_col: int = 12) -> None:
    worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[source_row].height

    for col in range(1, max_col + 1):
        source_cell = worksheet.cell(row=source_row, column=col)
        target_cell = worksheet.cell(row=target_row, column=col)

        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.font:
            target_cell.font = copy(source_cell.font)
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)
        if source_cell.border:
            target_cell.border = copy(source_cell.border)
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)


def ensure_capacity(worksheet, required_rows: int, start_row: int = 3, end_row: int = 8) -> None:
    existing_capacity = end_row - start_row + 1
    if required_rows <= existing_capacity:
        return

    additional_rows = required_rows - existing_capacity
    insert_at = end_row + 1
    worksheet.insert_rows(insert_at, additional_rows)

    template_rows = [3, 4, 5]
    for offset in range(additional_rows):
        target_row = insert_at + offset
        source_row = template_rows[offset % len(template_rows)]
        copy_row_format(worksheet, source_row, target_row)


def set_max_peak_formula(worksheet, row_index: int) -> None:
    worksheet.cell(
        row=row_index,
        column=4,
    ).value = (
        f'=IF(OR(B{row_index}="",F{row_index}=""),"",'
        f'F{row_index}+IFERROR(VLOOKUP(B{row_index},$K$3:$L$5,2,FALSE),0))'
    )


def unmerge_sample_unit_cells(worksheet) -> None:
    for merged_range in list(worksheet.merged_cells.ranges):
        if merged_range.min_col == 1 and 3 <= merged_range.min_row <= 8:
            worksheet.unmerge_cells(str(merged_range))


def merge_same_unit_id_cells(worksheet, start_row: int, end_row: int) -> None:
    if end_row < start_row:
        return

    merge_start = start_row
    current_value = worksheet.cell(row=start_row, column=1).value

    for row_index in range(start_row + 1, end_row + 1):
        next_value = worksheet.cell(row=row_index, column=1).value
        if next_value == current_value:
            continue

        if current_value not in (None, "") and row_index - merge_start > 1:
            worksheet.merge_cells(
                start_row=merge_start,
                start_column=1,
                end_row=row_index - 1,
                end_column=1,
            )

        merge_start = row_index
        current_value = next_value

    if current_value not in (None, "") and end_row - merge_start >= 1:
        worksheet.merge_cells(
            start_row=merge_start,
            start_column=1,
            end_row=end_row,
            end_column=1,
        )


def write_summary_excel(rows: list[dict], template_path: Path, output_file: Path) -> None:
    copy_template_file(template_path, output_file)

    workbook = load_workbook(output_file)
    worksheet = workbook["Sheet1"]

    excel_rows = sort_rows_for_excel(rows)
    ensure_capacity(worksheet, len(excel_rows))
    unmerge_sample_unit_cells(worksheet)

    # Clear the existing sample data area while preserving the template headers and factor table.
    last_data_row = max(8, 2 + len(excel_rows))
    for row_index in range(3, last_data_row + 1):
        for column in (1, 2, 3, 4, 5, 6):
            worksheet.cell(row=row_index, column=column).value = None

    for row_index, row in enumerate(excel_rows, start=3):
        worksheet.cell(row=row_index, column=1).value = row.get("unit_id")
        worksheet.cell(row=row_index, column=2).value = row.get("frequency_mhz")
        worksheet.cell(row=row_index, column=3).value = row.get("trp_dbm")
        worksheet.cell(row=row_index, column=6).value = row.get("max_peak_vpol_dbm")
        set_max_peak_formula(worksheet, row_index)

    merge_same_unit_id_cells(worksheet, 3, 2 + len(excel_rows))

    workbook.save(output_file)

def open_output_folder(path: Path):
    try:
        folder = path.parent.resolve()
        os.startfile(folder)
    except Exception as e:
        print(f"Could not open folder: {e}")


def main():
    root_dir = Path(DEFAULT_ROOT_DIR)
    template_path = DEFAULT_TEMPLATE_PATH
    excel_output = DEFAULT_EXCEL_OUTPUT

    if not root_dir.exists() or not root_dir.is_dir():
        print(f"Error: '{root_dir}' is not a valid directory")
        sys.exit(1)

    if not template_path.exists():
        print(f"Error: template file '{template_path}' was not found")
        sys.exit(1)

    print(f"Using root directory: {root_dir}")

    rows = scan_root(root_dir)
    write_summary_excel(rows, template_path, excel_output)

    print(f"Done. Parsed {len(rows)} files.")
    print(f"Excel summary: {excel_output.resolve()}")

    open_output_folder(excel_output)


if __name__ == "__main__":
    main()
